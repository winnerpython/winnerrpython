import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook(filename='/Users/winner/Desktop/reviews.xlsx')
ws = wb.active

# Launch the web driver (ensure you have installed the appropriate browser driver)
driver = webdriver.Chrome('/Users/winner/Desktop/chromedriver')

# Iterate over the rows in the Excel file
for row in ws.iter_rows(min_row=2):
    apple_id = row[0].value  # Apple ID from Excel
    password = row[1].value  # Password from Excel
    rating = row[2].value  # Rating value from Excel (e.g., 5)
    review_text = row[3].value  # Review text from Excel
    
    # Open the App Store page for your app
    driver.get('https://apps.apple.com/us/app/crypto-com-buy-btc-eth-shib/id1262148500')

    # Wait for the page to load
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'we-star-rating-star')))
    
    # Log in with Apple ID
    login_button = driver.find_element(By.CLASS_NAME, 'signin')
    login_button.click()
    
    apple_id_field = driver.find_element(By.ID, 'account_name_text_field')
    apple_id_field.send_keys(apple_id)
    
    password_field = driver.find_element(By.ID, 'password_text_field')
    password_field.send_keys(password)
    
    sign_in_button = driver.find_element(By.ID, 'sign-in')
    sign_in_button.click()
    
    # Wait for the login process to complete
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'we-star-rating-star')))
    
    # Click on the rating stars to select the desired rating
    rating_stars = driver.find_elements(By.CLASS_NAME, 'we-star-rating-star')
    rating_stars[rating - 1].click()  # Select the corresponding star (rating - 1)
    
    # Find the review text input field and enter your review text
    review_text_field = driver.find_element(By.CLASS_NAME, 'we-customer-review__textarea')
    review_text_field.send_keys(review_text)
    
    # Submit the review
    submit_button = driver.find_element(By.CLASS_NAME, 'we-customer-review__submit')
    submit_button.click()
    
    # Wait for the specified interval before leaving the next comment
    time.sleep(600)  # 10 minutes (600 seconds)

# Close the web driver
driver.quit()